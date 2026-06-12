# coding: utf-8
"""生産管理指標(YYYY年度-検査) CSV/Excel → inspection_management 取込（CLI・ファイル監視共用）"""
from __future__ import annotations

import csv
import hashlib
import json
import logging
import re
import warnings
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Iterable, Iterator, Optional

logger = logging.getLogger(__name__)

INSPECTION_DEFECT_PROCESS_CD = "KT09"
EXTERNAL_SYNC_KEY_PREFIX = "insp_excel:"
REMARKS_PREFIX_EXCEL = "EXCEL_SYNC"
REMARKS_PREFIX_CSV = "CSV_IMPORT"

DATA_SOURCE_MES = "mes"
DATA_SOURCE_EXCEL = "excel"
DATA_SOURCE_CSV = "csv"
VALID_DATA_SOURCES = frozenset({DATA_SOURCE_MES, DATA_SOURCE_EXCEL, DATA_SOURCE_CSV})


def data_source_from_remarks_prefix(remarks_prefix: str) -> str:
    if remarks_prefix == REMARKS_PREFIX_CSV:
        return DATA_SOURCE_CSV
    if remarks_prefix == REMARKS_PREFIX_EXCEL:
        return DATA_SOURCE_EXCEL
    return DATA_SOURCE_MES


def resolve_data_source(
    data_source: Optional[str],
    remarks: Optional[str] = None,
    external_sync_key: Optional[str] = None,
) -> str:
    """DB 列未設定の旧行向けに remarks / external_sync_key から推定。"""
    ds = (data_source or "").strip().lower()
    if ds in VALID_DATA_SOURCES:
        return ds
    remarks_s = (remarks or "").strip()
    if remarks_s.startswith(f"{REMARKS_PREFIX_EXCEL}:"):
        return DATA_SOURCE_EXCEL
    if remarks_s.startswith(f"{REMARKS_PREFIX_CSV}:"):
        return DATA_SOURCE_CSV
    if external_sync_key:
        return DATA_SOURCE_EXCEL
    return DATA_SOURCE_MES

COL_PRODUCT_CD = "社内品番"
COL_DAY_SHORT = "日付"
COL_INSPECTOR = "作業者"
COL_PRODUCT_NAME = "品名"
COL_ACTUAL_QTY = "組付合計"
COL_INSPECTION_TOTAL = "検査総数"
COL_DEFECT_TOTAL = "不良合計"
COL_SHIFT = "シフト"
COL_BREAK = "休憩"
COL_STOP = "停止(段替、待ち等)時間"
COL_WORK = "作業時間"
COL_RESULT_DAY = "実績日"

DEFECT_HEADERS = [
    "加工キズ",
    "油タレ",
    "曲げ不良",
    "カ他",
    "メッキ後キズ",
    "モヤ/カブリ",
    "ニッケル",
    "接触",
    "メ他",
    "溶接不良",
    "サビ",
    "生地不良",
    "外注メッキ不良",
    "外注溶接不良",
    "W検査　廃棄",
]

LOGICAL_COLUMNS = [
    COL_PRODUCT_CD,
    COL_DAY_SHORT,
    COL_INSPECTOR,
    COL_PRODUCT_NAME,
    COL_ACTUAL_QTY,
    COL_INSPECTION_TOTAL,
    COL_DEFECT_TOTAL,
    COL_SHIFT,
    COL_BREAK,
    COL_STOP,
    COL_WORK,
    COL_RESULT_DAY,
    *DEFECT_HEADERS,
]

FISCAL_YEAR_RE = re.compile(r"(\d{4})年度")
PRODUCT_CD_RE = re.compile(r"^\d{4,6}$")


def _norm_header(s: str) -> str:
    return (s or "").strip().replace("\u3000", " ").strip()


def _norm_name(s: str) -> str:
    return _norm_header(s).replace(" ", "").lower()


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y/%m/%d") if value.hour == 0 and value.minute == 0 else value.isoformat()
    if isinstance(value, date) and not isinstance(value, datetime):
        return value.strftime("%Y/%m/%d")
    return str(value).strip()


def _parse_int(value: Any) -> int:
    if value is None:
        return 0
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return max(0, value)
    if isinstance(value, float):
        if value != value:
            return 0
        return max(0, int(value))
    s = _cell_str(value).replace(",", "").replace("，", "")
    if not s or s in ("-", "—", "#N/A"):
        return 0
    try:
        return max(0, int(float(s)))
    except (TypeError, ValueError):
        return 0


def _parse_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        n = float(value)
        return n if n == n else None
    s = _cell_str(value).replace(",", "").replace("，", "").replace("%", "")
    if not s or s in ("-", "—", "#N/A"):
        return None
    try:
        n = float(s)
        return n if n == n else None
    except (TypeError, ValueError):
        return None


def _parse_date(value: Any, fiscal_year: int | None) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = _cell_str(value)
    if not s or s in ("#N/A", "-", "—"):
        return None
    for fmt in ("%Y/%m/%d", "%Y-%m-%d"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            pass
    m = re.match(r"^(\d{4})[/-](\d{1,2})[/-](\d{1,2})$", s)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    m = re.match(r"^(\d{1,2})[/-](\d{1,2})$", s)
    if m and fiscal_year:
        try:
            return date(fiscal_year, int(m.group(1)), int(m.group(2)))
        except ValueError:
            return None
    return None


def fiscal_year_from_source(source: str) -> int | None:
    m = FISCAL_YEAR_RE.search((source or "").replace("\uFF08", "(").replace("\uFF09", ")"))
    if not m:
        return None
    return int(m.group(1))


def is_data_row(row: dict[str, str]) -> bool:
    cd = (row.get(COL_PRODUCT_CD) or "").strip()
    if not cd or cd.startswith("#") or cd == "集計":
        return False
    if not PRODUCT_CD_RE.match(cd):
        return False
    if not (row.get(COL_INSPECTOR) or "").strip():
        return False
    qty = _parse_int(row.get(COL_ACTUAL_QTY)) or _parse_int(row.get(COL_INSPECTION_TOTAL))
    return qty > 0


@dataclass
class ParsedInspectionRow:
    source_line: int
    production_day: date
    production_month: date
    product_cd: str
    product_name: str
    inspector_name: str
    inspector_user_id: int | None
    actual_qty: int
    defect_qty: int
    mes_defect_by_item: dict[str, int]
    mes_net_production_sec: int | None
    mes_paused_accum_sec: int | None
    mes_production_started_at: datetime | None
    mes_production_ended_at: datetime | None
    external_sync_key: str
    data_source: str
    remarks: str
    production_sequence: int = 0
    warnings: list[str] = field(default_factory=list)


@dataclass
class ImportSyncResult:
    parsed: int = 0
    inserted: int = 0
    skipped_duplicate: int = 0
    skipped_unmapped: int = 0
    errors: list[str] = field(default_factory=list)
    unmapped_inspectors: dict[str, int] = field(default_factory=dict)


def build_header_index(headers: Iterable[Any]) -> dict[str, int]:
    idx: dict[str, int] = {}
    for i, h in enumerate(headers):
        key = _norm_header(_cell_str(h))
        if key and key not in idx:
            idx[key] = i
    return idx


def row_values_to_logical(header_index: dict[str, int], values: tuple[Any, ...]) -> dict[str, str]:
    out: dict[str, str] = {}
    for col in LOGICAL_COLUMNS:
        pos = header_index.get(_norm_header(col))
        if pos is None or pos >= len(values):
            out[col] = ""
        else:
            out[col] = _cell_str(values[pos])
    return out


def build_defect_map(cursor) -> dict[str, str]:
    cursor.execute(
        """
        SELECT defect_cd, defect_name
        FROM process_defect_items
        WHERE detection_process_cd = %s AND status = 'active'
        """,
        (INSPECTION_DEFECT_PROCESS_CD,),
    )
    out: dict[str, str] = {}
    for defect_cd, defect_name in cursor.fetchall():
        key = _norm_name(defect_name)
        if key:
            out[key] = str(defect_cd).strip()
    return out


def load_user_rows(cursor) -> list[tuple[int, str, str]]:
    cursor.execute("SELECT id, username, full_name FROM users")
    return [(int(uid), str(username or ""), str(full_name or "")) for uid, username, full_name in cursor.fetchall()]


def resolve_inspector_user_id(
    inspector_name: str,
    user_rows: list[tuple[int, str, str]],
    worker_overrides: dict[str, int] | None = None,
) -> int | None:
    """作業者名 → users.id（full_name 完全一致を最優先）"""
    name = (inspector_name or "").strip()
    if not name:
        return None
    overrides = worker_overrides or {}
    if name in overrides:
        return overrides[name]

    # 1) full_name 完全一致（前後空白のみ除去）
    for uid, _username, full_name in user_rows:
        if (full_name or "").strip() == name:
            return uid

    key = _norm_name(name)
    if not key:
        return None

    # 2) 正規化完全一致（full_name / username）
    for uid, username, full_name in user_rows:
        for raw in (full_name, username):
            if _norm_name(raw) == key:
                return uid

    # 3) 短名：full_name が作業者名で始まる / 含む
    for uid, username, full_name in user_rows:
        for raw in (full_name, username):
            n = _norm_name(raw)
            if not n:
                continue
            if n.startswith(key) or key in n:
                return uid
    return None


def map_defects(row: dict[str, str], defect_name_map: dict[str, str]) -> tuple[dict[str, int], list[str]]:
    defects: dict[str, int] = {}
    warnings: list[str] = []
    for header in DEFECT_HEADERS:
        qty = _parse_int(row.get(header))
        if qty <= 0:
            continue
        defect_cd = defect_name_map.get(_norm_name(header))
        if not defect_cd:
            defect_cd = f"csv:{header.strip()}"
            warnings.append(f"不良項目未マスタ: {header} → {defect_cd}")
        defects[defect_cd] = defects.get(defect_cd, 0) + qty
    return defects, warnings


def compute_time_fields(
    row: dict[str, str], production_day: date
) -> tuple[int | None, int | None, datetime | None, datetime | None]:
    shift_h = _parse_float(row.get(COL_SHIFT))
    break_h = _parse_float(row.get(COL_BREAK)) or 0.0
    stop_h = _parse_float(row.get(COL_STOP)) or 0.0
    work_h = _parse_float(row.get(COL_WORK))
    if work_h is None or work_h <= 0:
        return None, None, None, None
    if shift_h is None or shift_h <= 0:
        shift_h = work_h + break_h + stop_h
    pause_h = max(0.0, shift_h - work_h)
    if break_h + stop_h > pause_h:
        pause_h = break_h + stop_h
    net_sec = max(0, round(work_h * 3600))
    pause_sec = max(0, round(pause_h * 3600))
    started = datetime.combine(production_day, time(8, 0, 0))
    ended = started + timedelta(seconds=max(0, round(shift_h * 3600)))
    return net_sec, pause_sec, started, ended


def make_external_sync_key(
    production_day: date,
    product_cd: str,
    inspector_user_id: int | None,
    actual_qty: int,
    defect_qty: int,
    mes_defect_by_item: dict[str, int],
) -> str:
    payload = "|".join(
        [
            production_day.isoformat(),
            product_cd.strip(),
            str(inspector_user_id or ""),
            str(actual_qty),
            str(defect_qty),
            json.dumps(mes_defect_by_item or {}, sort_keys=True, ensure_ascii=False),
        ]
    )
    digest = hashlib.sha256(payload.encode("utf-8")).hexdigest()[:40]
    return f"{EXTERNAL_SYNC_KEY_PREFIX}{digest}"


def parse_logical_row(
    logical: dict[str, str],
    *,
    source_line: int,
    source_label: str,
    remarks_prefix: str,
    fiscal_year: int | None,
    defect_name_map: dict[str, str],
    user_rows: list[tuple[int, str, str]],
    worker_overrides: dict[str, int] | None = None,
    skip_unmapped_inspector: bool = False,
) -> ParsedInspectionRow | None:
    if not is_data_row(logical):
        return None

    prod_day = _parse_date(logical.get(COL_RESULT_DAY), fiscal_year)
    if prod_day is None:
        prod_day = _parse_date(logical.get(COL_DAY_SHORT), fiscal_year)
    if prod_day is None:
        return None

    inspector_name = (logical.get(COL_INSPECTOR) or "").strip()
    inspector_id = resolve_inspector_user_id(inspector_name, user_rows, worker_overrides)
    if inspector_id is None and skip_unmapped_inspector:
        return None

    actual_qty = _parse_int(logical.get(COL_ACTUAL_QTY)) or _parse_int(logical.get(COL_INSPECTION_TOTAL))
    defect_qty_csv = _parse_int(logical.get(COL_DEFECT_TOTAL))
    defects, defect_warnings = map_defects(logical, defect_name_map)
    defect_qty = defect_qty_csv if defect_qty_csv > 0 else sum(defects.values())
    net_sec, pause_sec, started, ended = compute_time_fields(logical, prod_day)
    product_name = (logical.get(COL_PRODUCT_NAME) or "").strip() or logical.get(COL_PRODUCT_CD, "").strip()

    row_warnings: list[str] = list(defect_warnings)
    if inspector_id is None:
        row_warnings.append(f"作業者未匹配 users.full_name: {inspector_name}")

    sync_key = make_external_sync_key(
        prod_day, logical.get(COL_PRODUCT_CD, "").strip(), inspector_id, actual_qty, defect_qty, defects
    )
    return ParsedInspectionRow(
        source_line=source_line,
        production_day=prod_day,
        production_month=prod_day.replace(day=1),
        product_cd=logical.get(COL_PRODUCT_CD, "").strip(),
        product_name=product_name,
        inspector_name=inspector_name,
        inspector_user_id=inspector_id,
        actual_qty=actual_qty,
        defect_qty=defect_qty,
        mes_defect_by_item=defects,
        mes_net_production_sec=net_sec,
        mes_paused_accum_sec=pause_sec,
        mes_production_started_at=started,
        mes_production_ended_at=ended,
        external_sync_key=sync_key,
        data_source=data_source_from_remarks_prefix(remarks_prefix),
        remarks=f"{remarks_prefix}:{source_label}:L{source_line}",
        warnings=row_warnings,
    )


def iter_csv_logical_rows(csv_path: Path) -> Iterator[tuple[int, dict[str, str]]]:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            return
        header_index = build_header_index(reader.fieldnames)
        for line_no, raw in enumerate(reader, start=2):
            values = tuple(raw.get(h, "") for h in reader.fieldnames)
            yield line_no, row_values_to_logical(header_index, values)


def iter_excel_logical_rows(filepath: str) -> Iterator[tuple[int, dict[str, str]]]:
    import openpyxl

    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="Print area cannot be set to Defined name")
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    try:
        sheet = wb.active
        if sheet is None or (sheet.max_row or 0) <= 1:
            return
        rows_iter = sheet.iter_rows(min_row=1, values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return
        header_index = build_header_index(header_row)
        for line_no, values in enumerate(rows_iter, start=2):
            if not values:
                continue
            yield line_no, row_values_to_logical(header_index, values)
    finally:
        try:
            wb.close()
        except Exception:
            pass


def parse_rows_from_source(
    source_path: str,
    *,
    defect_name_map: dict[str, str],
    user_rows: list[tuple[int, str, str]],
    worker_overrides: dict[str, int] | None = None,
    skip_unmapped_inspector: bool = False,
    remarks_prefix: str = REMARKS_PREFIX_EXCEL,
) -> tuple[list[ParsedInspectionRow], list[str]]:
    path = Path(source_path)
    source_label = path.name
    fiscal_year = fiscal_year_from_source(source_label)
    parsed: list[ParsedInspectionRow] = []
    errors: list[str] = []

    suffix = path.suffix.lower()
    if suffix == ".csv":
        row_iter = iter_csv_logical_rows(path)
    elif suffix in (".xlsx", ".xlsm"):
        row_iter = iter_excel_logical_rows(str(path))
    else:
        raise ValueError(f"未対応の拡張子: {suffix}")

    for line_no, logical in row_iter:
        if not is_data_row(logical):
            continue
        row = parse_logical_row(
            logical,
            source_line=line_no,
            source_label=source_label,
            remarks_prefix=remarks_prefix,
            fiscal_year=fiscal_year,
            defect_name_map=defect_name_map,
            user_rows=user_rows,
            worker_overrides=worker_overrides,
            skip_unmapped_inspector=skip_unmapped_inspector,
        )
        if row is None:
            insp = (logical.get(COL_INSPECTOR) or "").strip()
            if skip_unmapped_inspector and insp:
                errors.append(f"行 {line_no}: 作業者未匹配のためスキップ ({insp})")
            else:
                errors.append(f"行 {line_no}: 解析不可")
            continue
        parsed.append(row)
    return parsed, errors


def _table_has_column(cursor, table: str, column: str) -> bool:
    cursor.execute(
        """
        SELECT 1 FROM information_schema.columns
        WHERE table_schema = DATABASE() AND table_name = %s AND column_name = %s
        LIMIT 1
        """,
        (table, column),
    )
    return cursor.fetchone() is not None


def _defect_json_for_fingerprint(raw: Any) -> str:
    if raw is None:
        return "{}"
    if isinstance(raw, str):
        try:
            data = json.loads(raw) if raw.strip() else {}
        except json.JSONDecodeError:
            return raw
    elif isinstance(raw, dict):
        data = raw
    else:
        return "{}"
    if not isinstance(data, dict):
        return "{}"
    cleaned = {str(k): int(v) for k, v in data.items() if v is not None and int(v) > 0}
    return json.dumps(cleaned, sort_keys=True, ensure_ascii=False)


def business_fingerprint(
    production_day: date,
    product_cd: str,
    inspector_user_id: int | None,
    actual_qty: int,
    defect_qty: int,
    mes_defect_by_item: Any = None,
) -> tuple:
    return (
        production_day,
        product_cd.strip(),
        inspector_user_id,
        int(actual_qty),
        int(defect_qty),
    )


def _load_existing_business_fingerprints(cursor, days: Iterable[date]) -> set[tuple]:
    day_list = list(days)
    if not day_list:
        return set()
    placeholders = ",".join(["%s"] * len(day_list))
    cursor.execute(
        f"""
        SELECT production_day, product_cd, mes_inspector_user_id,
               actual_production_quantity, defect_qty, mes_defect_by_item
        FROM inspection_management
        WHERE production_day IN ({placeholders})
        """,
        day_list,
    )
    fps: set[tuple] = set()
    for prod_day, product_cd, insp_id, actual_qty, defect_qty, defects in cursor.fetchall():
        fps.add(
            business_fingerprint(
                prod_day, str(product_cd or ""), insp_id, int(actual_qty or 0), int(defect_qty or 0), defects
            )
        )
    return fps


def _load_existing_sync_keys(cursor, keys: list[str]) -> set[str]:
    if not keys:
        return set()
    existing: set[str] = set()
    chunk = 500
    for i in range(0, len(keys), chunk):
        batch = keys[i : i + chunk]
        placeholders = ",".join(["%s"] * len(batch))
        if _table_has_column(cursor, "inspection_management", "external_sync_key"):
            cursor.execute(
                f"SELECT external_sync_key FROM inspection_management "
                f"WHERE external_sync_key IN ({placeholders})",
                batch,
            )
            existing.update(str(r[0]) for r in cursor.fetchall() if r[0])
    return existing


def _next_sequences(cursor, days: Iterable[date]) -> dict[date, int]:
    seq: dict[date, int] = {}
    for d in days:
        cursor.execute(
            "SELECT COALESCE(MAX(production_sequence), 0) FROM inspection_management WHERE production_day = %s",
            (d,),
        )
        seq[d] = int(cursor.fetchone()[0] or 0)
    return seq


def assign_production_sequences(rows: list[ParsedInspectionRow], cursor) -> None:
    days = {r.production_day for r in rows}
    counters = _next_sequences(cursor, days)
    for row in rows:
        counters[row.production_day] += 1
        row.production_sequence = counters[row.production_day]


def sync_parsed_rows_to_db(
    cursor,
    rows: list[ParsedInspectionRow],
    *,
    dry_run: bool = False,
) -> ImportSyncResult:
    result = ImportSyncResult(parsed=len(rows))
    if not rows:
        return result

    keys = [r.external_sync_key for r in rows]
    existing_keys = _load_existing_sync_keys(cursor, keys)
    new_rows = [r for r in rows if r.external_sync_key not in existing_keys]

    biz_fps = _load_existing_business_fingerprints(cursor, {r.production_day for r in new_rows})
    deduped: list[ParsedInspectionRow] = []
    for row in new_rows:
        fp = business_fingerprint(
            row.production_day,
            row.product_cd,
            row.inspector_user_id,
            row.actual_qty,
            row.defect_qty,
            row.mes_defect_by_item,
        )
        if fp in biz_fps:
            continue
        deduped.append(row)
        biz_fps.add(fp)
    result.skipped_duplicate = len(rows) - len(deduped)
    new_rows = deduped

    for r in rows:
        if r.inspector_user_id is None:
            result.unmapped_inspectors[r.inspector_name] = result.unmapped_inspectors.get(r.inspector_name, 0) + 1

    if not new_rows or dry_run:
        result.inserted = len(new_rows) if dry_run else 0
        return result

    assign_production_sequences(new_rows, cursor)
    has_data_source_col = _table_has_column(cursor, "inspection_management", "data_source")
    has_sync_col = _table_has_column(cursor, "inspection_management", "external_sync_key")

    insert_cols = [
        "production_month",
        "production_day",
        "production_sequence",
        "product_cd",
        "product_name",
        "actual_production_quantity",
        "defect_qty",
        "mes_defect_by_item",
        "production_completed_check",
        "mes_production_started_at",
        "mes_production_ended_at",
        "mes_net_production_sec",
        "mes_paused_accum_sec",
        "mes_production_is_paused",
        "mes_inspector_user_id",
        "remarks",
    ]
    if has_sync_col:
        insert_cols.append("external_sync_key")
    if has_data_source_col:
        insert_cols.append("data_source")
    placeholders = ", ".join(["%s"] * len(insert_cols))
    sql = f"INSERT INTO inspection_management ({', '.join(insert_cols)}) VALUES ({placeholders})"

    batch_params = []
    for row in new_rows:
        defect_json = json.dumps(row.mes_defect_by_item, ensure_ascii=False) if row.mes_defect_by_item else None
        values: list[Any] = [
            row.production_month,
            row.production_day,
            row.production_sequence,
            row.product_cd,
            row.product_name,
            row.actual_qty,
            row.defect_qty,
            defect_json,
            1,
            row.mes_production_started_at,
            row.mes_production_ended_at,
            row.mes_net_production_sec,
            row.mes_paused_accum_sec,
            0,
            row.inspector_user_id,
            row.remarks,
        ]
        if has_sync_col:
            values.append(row.external_sync_key)
        if has_data_source_col:
            values.append(row.data_source)
        batch_params.append(tuple(values))

    cursor.executemany(sql, batch_params)
    result.inserted = len(new_rows)
    return result


def delete_rows_by_remarks_prefix(cursor, prefix: str, dry_run: bool = False) -> int:
    pattern = f"{prefix}%"
    cursor.execute(
        "SELECT COUNT(1) FROM inspection_management WHERE remarks LIKE %s",
        (pattern,),
    )
    count = int(cursor.fetchone()[0])
    if count and not dry_run:
        cursor.execute("DELETE FROM inspection_management WHERE remarks LIKE %s", (pattern,))
    return count


def sync_inspection_source_file(
    source_path: str,
    connection_factory,
    *,
    dry_run: bool = False,
    skip_unmapped_inspector: bool = False,
    worker_overrides: dict[str, int] | None = None,
    remarks_prefix: str = REMARKS_PREFIX_EXCEL,
) -> ImportSyncResult:
    """CSV/Excel 1 ファイルを inspection_management に增量同期（内容ハッシュで重複排除）"""
    conn = connection_factory()
    conn.autocommit = False
    cursor = conn.cursor()
    try:
        defect_map = build_defect_map(cursor)
        user_rows = load_user_rows(cursor)
        rows, parse_errors = parse_rows_from_source(
            source_path,
            defect_name_map=defect_map,
            user_rows=user_rows,
            worker_overrides=worker_overrides,
            skip_unmapped_inspector=skip_unmapped_inspector,
            remarks_prefix=remarks_prefix,
        )
        result = sync_parsed_rows_to_db(cursor, rows, dry_run=dry_run)
        result.errors.extend(parse_errors)
        if not dry_run:
            conn.commit()
        logger.info(
            "inspection_management 同期完了: parsed=%s inserted=%s dup_skip=%s path=%s",
            result.parsed,
            result.inserted,
            result.skipped_duplicate,
            source_path,
        )
        return result
    except Exception:
        conn.rollback()
        raise
    finally:
        cursor.close()
        conn.close()
