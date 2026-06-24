# coding: utf-8
"""生産管理指標(YYYY年度-溶接) CSV/Excel「入力」→ welding_management 取込（CLI・ファイル監視共用）"""
from __future__ import annotations

import csv
import hashlib
import json
import logging
import re
import warnings
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Iterable, Iterator, Optional

logger = logging.getLogger(__name__)

WELDING_INPUT_SHEET = "入力"
EXTERNAL_SYNC_KEY_PREFIX = "weld_excel:"
REMARKS_PREFIX_EXCEL = "EXCEL_SYNC"
REMARKS_PREFIX_CSV = "CSV_IMPORT"

DATA_SOURCE_MES = "mes"
DATA_SOURCE_EXCEL = "excel"
DATA_SOURCE_CSV = "csv"
VALID_DATA_SOURCES = frozenset({DATA_SOURCE_MES, DATA_SOURCE_EXCEL, DATA_SOURCE_CSV})

COL_PRODUCT_CD = "社内品番"
COL_DAY_SHORT = "日付"
COL_OPERATOR = "作業者"
COL_LINE = "ライン"
COL_PRODUCT_NAME = "品名"
COL_ACTUAL_QTY = "生産数"
COL_DEFECT_TOTAL = "不良数"
COL_SHIFT = "シフト"
COL_OVERTIME = "残業"
COL_BREAK = "休憩"
COL_WORK = "作業時間"
COL_AVAILABLE = "稼働可能時間"

PAUSE_CATEGORY_HEADERS = [
    "交換（チップ、ワイヤ）",
    "停止（運搬）",
    "清掃",
    "準備",
    "その他",
    "調整",
    "修理",
]

LOGICAL_COLUMNS = [
    COL_PRODUCT_CD,
    COL_DAY_SHORT,
    COL_OPERATOR,
    COL_LINE,
    COL_PRODUCT_NAME,
    COL_ACTUAL_QTY,
    COL_DEFECT_TOTAL,
    COL_SHIFT,
    COL_OVERTIME,
    COL_BREAK,
    *PAUSE_CATEGORY_HEADERS,
    COL_AVAILABLE,
    COL_WORK,
]

# 入力表シート自動判定用（社内品番・日付・作業者が揃っていること）
WELDING_SHEET_HEADER_MARKERS = (COL_PRODUCT_CD, COL_DAY_SHORT, COL_OPERATOR)

FISCAL_YEAR_RE = re.compile(r"(\d{4})年度")
PRODUCT_CD_RE = re.compile(r"^\d{4,6}$")


def data_source_from_remarks_prefix(remarks_prefix: str) -> str:
    if remarks_prefix == REMARKS_PREFIX_CSV:
        return DATA_SOURCE_CSV
    if remarks_prefix == REMARKS_PREFIX_EXCEL:
        return DATA_SOURCE_EXCEL
    return DATA_SOURCE_MES


def resolve_data_source(
    data_source: Optional[str],
    remarks: Optional[str] = None,
    external_sync_key: Optional[str] = None,
) -> str:
    ds = (data_source or "").strip().lower()
    if ds in VALID_DATA_SOURCES:
        return ds
    remarks_s = (remarks or "").strip()
    if remarks_s.startswith(f"{REMARKS_PREFIX_EXCEL}:"):
        return DATA_SOURCE_EXCEL
    if remarks_s.startswith(f"{REMARKS_PREFIX_CSV}:"):
        return DATA_SOURCE_CSV
    if external_sync_key:
        return DATA_SOURCE_EXCEL
    return DATA_SOURCE_MES


def _norm_header(s: str) -> str:
    return (s or "").strip().replace("\u3000", " ").strip()


def _norm_name(s: str) -> str:
    return _norm_header(s).replace(" ", "").lower()


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y/%m/%d") if value.hour == 0 and value.minute == 0 else value.isoformat()
    if isinstance(value, date) and not isinstance(value, datetime):
        return value.strftime("%Y/%m/%d")
    return str(value).strip()


def _parse_int(value: Any) -> int:
    if value is None:
        return 0
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return max(0, value)
    if isinstance(value, float):
        if value != value:
            return 0
        return max(0, int(value))
    s = _cell_str(value).replace(",", "").replace("，", "")
    if not s or s in ("-", "—", "#N/A"):
        return 0
    try:
        return max(0, int(float(s)))
    except (TypeError, ValueError):
        return 0


def _parse_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        n = float(value)
        return n if n == n else None
    s = _cell_str(value).replace(",", "").replace("，", "").replace("%", "")
    if not s or s in ("-", "—", "#N/A"):
        return None
    try:
        n = float(s)
        return n if n == n else None
    except (TypeError, ValueError):
        return None


def _parse_date(value: Any, fiscal_year: int | None) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = _cell_str(value)
    if not s or s in ("#N/A", "-", "—"):
        return None
    for fmt in ("%Y/%m/%d", "%Y-%m-%d"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            pass
    m = re.match(r"^(\d{4})[/-](\d{1,2})[/-](\d{1,2})$", s)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    m = re.match(r"^(\d{1,2})[/-](\d{1,2})$", s)
    if m and fiscal_year:
        try:
            return date(fiscal_year, int(m.group(1)), int(m.group(2)))
        except ValueError:
            return None
    return None


def fiscal_year_from_source(source: str) -> int | None:
    m = FISCAL_YEAR_RE.search((source or "").replace("\uFF08", "(").replace("\uFF09", ")"))
    if not m:
        return None
    return int(m.group(1))


def is_data_row(row: dict[str, str]) -> bool:
    cd = (row.get(COL_PRODUCT_CD) or "").strip()
    if not cd or cd.startswith("#") or cd == "集計":
        return False
    if not PRODUCT_CD_RE.match(cd):
        return False
    if not (row.get(COL_OPERATOR) or "").strip():
        return False
    return _parse_int(row.get(COL_ACTUAL_QTY)) > 0


@dataclass
class ParsedWeldingRow:
    source_line: int
    production_day: date
    production_month: date
    product_cd: str
    product_name: str
    welding_machine: str
    operator_name: str
    operator_user_id: int | None
    actual_qty: int
    defect_qty: int
    mes_net_production_sec: int | None
    mes_paused_accum_sec: int | None
    mes_shift_sec: int | None
    mes_break_sec: int | None
    mes_stop_sec: int | None
    mes_production_started_at: datetime | None
    mes_production_ended_at: datetime | None
    external_sync_key: str
    data_source: str
    remarks: str
    production_sequence: int = 0
    warnings: list[str] = field(default_factory=list)


@dataclass
class ImportSyncResult:
    parsed: int = 0
    inserted: int = 0
    skipped_duplicate: int = 0
    skipped_unmapped: int = 0
    errors: list[str] = field(default_factory=list)
    unmapped_operators: dict[str, int] = field(default_factory=dict)


def build_header_index(headers: Iterable[Any]) -> dict[str, int]:
    idx: dict[str, int] = {}
    for i, h in enumerate(headers):
        key = _norm_header(_cell_str(h))
        if key and key not in idx:
            idx[key] = i
    return idx


def row_values_to_logical(header_index: dict[str, int], values: tuple[Any, ...]) -> dict[str, str]:
    out: dict[str, str] = {}
    for col in LOGICAL_COLUMNS:
        pos = header_index.get(_norm_header(col))
        if pos is None or pos >= len(values):
            out[col] = ""
        else:
            out[col] = _cell_str(values[pos])
    return out


def load_user_rows(cursor) -> list[tuple[int, str, str]]:
    cursor.execute("SELECT id, username, full_name FROM users")
    return [(int(uid), str(username or ""), str(full_name or "")) for uid, username, full_name in cursor.fetchall()]


def resolve_operator_user_id(
    operator_name: str,
    user_rows: list[tuple[int, str, str]],
    worker_overrides: dict[str, int] | None = None,
) -> int | None:
    """作業者名 → users.id（full_name 完全一致を最優先）"""
    name = (operator_name or "").strip()
    if not name:
        return None
    overrides = worker_overrides or {}
    if name in overrides:
        return overrides[name]

    for uid, _username, full_name in user_rows:
        if (full_name or "").strip() == name:
            return uid

    key = _norm_name(name)
    if not key:
        return None

    for uid, username, full_name in user_rows:
        for raw in (full_name, username):
            if _norm_name(raw) == key:
                return uid

    for uid, username, full_name in user_rows:
        for raw in (full_name, username):
            n = _norm_name(raw)
            if not n:
                continue
            if n.startswith(key) or key in n:
                return uid
    return None


def compute_time_fields(
    row: dict[str, str], production_day: date
) -> tuple[int | None, int | None, int | None, int | None, int | None, datetime | None, datetime | None]:
    shift_h = _parse_float(row.get(COL_SHIFT))
    overtime_h = _parse_float(row.get(COL_OVERTIME)) or 0.0
    if overtime_h < 0:
        overtime_h = 0.0
    break_h = _parse_float(row.get(COL_BREAK)) or 0.0
    stop_h = sum(_parse_float(row.get(h)) or 0.0 for h in PAUSE_CATEGORY_HEADERS)
    work_h = _parse_float(row.get(COL_WORK))
    if work_h is None or work_h <= 0:
        return None, None, None, None, None, None, None
    if shift_h is None or shift_h <= 0:
        available_h = _parse_float(row.get(COL_AVAILABLE))
        shift_h = available_h if available_h and available_h > 0 else work_h + break_h + stop_h
    effective_shift_h = shift_h + overtime_h
    pause_h = max(0.0, effective_shift_h - work_h)
    if break_h + stop_h > pause_h:
        pause_h = break_h + stop_h
    net_sec = max(0, round(work_h * 3600))
    pause_sec = max(0, round(pause_h * 3600))
    shift_sec = max(0, round(effective_shift_h * 3600))
    break_sec = max(0, round(break_h * 3600))
    stop_sec = max(0, round(stop_h * 3600))
    started = datetime.combine(production_day, time(8, 0, 0))
    # 終了時刻はシフト全体ではなく正味稼働（作業時間）に合わせる
    ended = started + timedelta(seconds=net_sec)
    return net_sec, pause_sec, shift_sec, break_sec, stop_sec, started, ended


def make_external_sync_key(
    production_day: date,
    product_cd: str,
    welding_machine: str,
    operator_user_id: int | None,
    actual_qty: int,
    defect_qty: int,
) -> str:
    payload = "|".join(
        [
            production_day.isoformat(),
            product_cd.strip(),
            (welding_machine or "").strip(),
            str(operator_user_id or ""),
            str(actual_qty),
            str(defect_qty),
        ]
    )
    digest = hashlib.sha256(payload.encode("utf-8")).hexdigest()[:40]
    return f"{EXTERNAL_SYNC_KEY_PREFIX}{digest}"


def parse_logical_row(
    logical: dict[str, str],
    *,
    source_line: int,
    source_label: str,
    remarks_prefix: str,
    fiscal_year: int | None,
    user_rows: list[tuple[int, str, str]],
    worker_overrides: dict[str, int] | None = None,
    skip_unmapped_operator: bool = False,
) -> ParsedWeldingRow | None:
    if not is_data_row(logical):
        return None

    prod_day = _parse_date(logical.get(COL_DAY_SHORT), fiscal_year)
    if prod_day is None:
        return None

    operator_name = (logical.get(COL_OPERATOR) or "").strip()
    operator_id = resolve_operator_user_id(operator_name, user_rows, worker_overrides)
    if operator_id is None and skip_unmapped_operator:
        return None

    actual_qty = _parse_int(logical.get(COL_ACTUAL_QTY))
    defect_qty = _parse_int(logical.get(COL_DEFECT_TOTAL))
    welding_machine = (logical.get(COL_LINE) or "").strip()
    net_sec, pause_sec, shift_sec, break_sec, stop_sec, started, ended = compute_time_fields(logical, prod_day)
    product_name = (logical.get(COL_PRODUCT_NAME) or "").strip() or logical.get(COL_PRODUCT_CD, "").strip()

    row_warnings: list[str] = []
    if operator_id is None:
        row_warnings.append(f"作業者未匹配 users.full_name: {operator_name}")

    sync_key = make_external_sync_key(
        prod_day,
        logical.get(COL_PRODUCT_CD, "").strip(),
        welding_machine,
        operator_id,
        actual_qty,
        defect_qty,
    )
    return ParsedWeldingRow(
        source_line=source_line,
        production_day=prod_day,
        production_month=prod_day.replace(day=1),
        product_cd=logical.get(COL_PRODUCT_CD, "").strip(),
        product_name=product_name,
        welding_machine=welding_machine,
        operator_name=operator_name,
        operator_user_id=operator_id,
        actual_qty=actual_qty,
        defect_qty=defect_qty,
        mes_net_production_sec=net_sec,
        mes_paused_accum_sec=pause_sec,
        mes_shift_sec=shift_sec,
        mes_break_sec=break_sec,
        mes_stop_sec=stop_sec,
        mes_production_started_at=started,
        mes_production_ended_at=ended,
        external_sync_key=sync_key,
        data_source=data_source_from_remarks_prefix(remarks_prefix),
        remarks=f"{remarks_prefix}:{source_label}:L{source_line}",
        warnings=row_warnings,
    )


def iter_csv_logical_rows(csv_path: Path) -> Iterator[tuple[int, dict[str, str]]]:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            return
        header_index = build_header_index(reader.fieldnames)
        for line_no, raw in enumerate(reader, start=2):
            values = tuple(raw.get(h, "") for h in reader.fieldnames)
            yield line_no, row_values_to_logical(header_index, values)


def iter_welding_excel_logical_rows(filepath: str, sheet_name: str = WELDING_INPUT_SHEET) -> Iterator[tuple[int, dict[str, str]]]:
    import openpyxl

    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="Print area cannot be set to Defined name")
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    try:
        resolved_name, sheet = resolve_welding_input_sheet(wb, preferred_name=sheet_name, filepath=filepath)
        if resolved_name != sheet_name:
            logger.info("溶接入力表: シート「%s」を使用（指定/既定「%s」なし）", resolved_name, sheet_name)
        if sheet is None or (sheet.max_row or 0) <= 1:
            return
        rows_iter = sheet.iter_rows(min_row=1, values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return
        header_index = build_header_index(header_row)
        for line_no, values in enumerate(rows_iter, start=2):
            if not values:
                continue
            yield line_no, row_values_to_logical(header_index, values)
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _sheet_header_row(sheet) -> tuple[Any, ...] | None:
    rows_iter = sheet.iter_rows(min_row=1, max_row=1, values_only=True)
    return next(rows_iter, None)


def _sheet_has_welding_input_headers(sheet) -> bool:
    header_row = _sheet_header_row(sheet)
    if not header_row:
        return False
    header_index = build_header_index(header_row)
    return all(_norm_header(col) in header_index for col in WELDING_SHEET_HEADER_MARKERS)


def resolve_welding_input_sheet(wb, *, preferred_name: str = WELDING_INPUT_SHEET, filepath: str = ""):
    """入力表シートを解決。優先名 → ヘッダ一致シート → active の順。"""
    if preferred_name in wb.sheetnames:
        sheet = wb[preferred_name]
        if _sheet_has_welding_input_headers(sheet):
            return preferred_name, sheet
        logger.warning(
            "シート「%s」は存在しますが入力表ヘッダがありません。他シートを探索します",
            preferred_name,
        )

    for name in wb.sheetnames:
        if name == preferred_name:
            continue
        sheet = wb[name]
        if _sheet_has_welding_input_headers(sheet):
            return name, sheet

    active = wb.active
    if active is not None and _sheet_has_welding_input_headers(active):
        return active.title, active

    available = ", ".join(wb.sheetnames)
    location = f"（{filepath}）" if filepath else ""
    raise ValueError(
        f"シート「{preferred_name}」がなく、入力表ヘッダ（社内品番/日付/作業者）を持つシートも見つかりません"
        f"{location}。利用可能: {available}"
    )


def parse_rows_from_source(
    source_path: str,
    *,
    user_rows: list[tuple[int, str, str]],
    worker_overrides: dict[str, int] | None = None,
    skip_unmapped_operator: bool = False,
    remarks_prefix: str = REMARKS_PREFIX_EXCEL,
) -> tuple[list[ParsedWeldingRow], list[str]]:
    path = Path(source_path)
    source_label = path.name
    fiscal_year = fiscal_year_from_source(source_label)
    parsed: list[ParsedWeldingRow] = []
    errors: list[str] = []

    suffix = path.suffix.lower()
    if suffix == ".csv":
        row_iter = iter_csv_logical_rows(path)
    elif suffix in (".xlsx", ".xlsm"):
        row_iter = iter_welding_excel_logical_rows(str(path))
    else:
        raise ValueError(f"未対応の拡張子: {suffix}")

    for line_no, logical in row_iter:
        if not is_data_row(logical):
            continue
        row = parse_logical_row(
            logical,
            source_line=line_no,
            source_label=source_label,
            remarks_prefix=remarks_prefix,
            fiscal_year=fiscal_year,
            user_rows=user_rows,
            worker_overrides=worker_overrides,
            skip_unmapped_operator=skip_unmapped_operator,
        )
        if row is None:
            op = (logical.get(COL_OPERATOR) or "").strip()
            if skip_unmapped_operator and op:
                errors.append(f"行 {line_no}: 作業者未匹配のためスキップ ({op})")
            else:
                errors.append(f"行 {line_no}: 解析不可")
            continue
        parsed.append(row)
    return parsed, errors


def _table_has_column(cursor, table: str, column: str) -> bool:
    cursor.execute(
        """
        SELECT 1 FROM information_schema.columns
        WHERE table_schema = DATABASE() AND table_name = %s AND column_name = %s
        LIMIT 1
        """,
        (table, column),
    )
    return cursor.fetchone() is not None


def business_fingerprint(
    production_day: date,
    product_cd: str,
    welding_machine: str,
    operator_user_id: int | None,
    actual_qty: int,
    defect_qty: int,
) -> tuple:
    return (
        production_day,
        product_cd.strip(),
        (welding_machine or "").strip(),
        operator_user_id,
        int(actual_qty),
        int(defect_qty),
    )


def _load_existing_business_fingerprints(cursor, days: Iterable[date]) -> set[tuple]:
    day_list = list(days)
    if not day_list:
        return set()
    placeholders = ",".join(["%s"] * len(day_list))
    has_machine = _table_has_column(cursor, "welding_management", "welding_machine")
    if has_machine:
        cursor.execute(
            f"""
            SELECT production_day, product_cd, welding_machine, mes_operator_user_id,
                   actual_production_quantity, defect_qty
            FROM welding_management
            WHERE production_day IN ({placeholders})
            """,
            day_list,
        )
        rows = cursor.fetchall()
        fps: set[tuple] = set()
        for prod_day, product_cd, machine, op_id, actual_qty, defect_qty in rows:
            fps.add(
                business_fingerprint(
                    prod_day,
                    str(product_cd or ""),
                    str(machine or ""),
                    op_id,
                    int(actual_qty or 0),
                    int(defect_qty or 0),
                )
            )
        return fps

    cursor.execute(
        f"""
        SELECT production_day, product_cd, mes_operator_user_id,
               actual_production_quantity, defect_qty
        FROM welding_management
        WHERE production_day IN ({placeholders})
        """,
        day_list,
    )
    fps = set()
    for prod_day, product_cd, op_id, actual_qty, defect_qty in cursor.fetchall():
        fps.add(
            business_fingerprint(
                prod_day,
                str(product_cd or ""),
                "",
                op_id,
                int(actual_qty or 0),
                int(defect_qty or 0),
            )
        )
    return fps


def _load_existing_sync_keys(cursor, keys: list[str]) -> set[str]:
    if not keys:
        return set()
    existing: set[str] = set()
    chunk = 500
    for i in range(0, len(keys), chunk):
        batch = keys[i : i + chunk]
        placeholders = ",".join(["%s"] * len(batch))
        if _table_has_column(cursor, "welding_management", "external_sync_key"):
            cursor.execute(
                f"SELECT external_sync_key FROM welding_management "
                f"WHERE external_sync_key IN ({placeholders})",
                batch,
            )
            existing.update(str(r[0]) for r in cursor.fetchall() if r[0])
    return existing


def _next_sequences(cursor, days: Iterable[date]) -> dict[date, int]:
    seq: dict[date, int] = {}
    for d in days:
        cursor.execute(
            "SELECT COALESCE(MAX(production_sequence), 0) FROM welding_management WHERE production_day = %s",
            (d,),
        )
        seq[d] = int(cursor.fetchone()[0] or 0)
    return seq


def assign_production_sequences(rows: list[ParsedWeldingRow], cursor) -> None:
    days = {r.production_day for r in rows}
    counters = _next_sequences(cursor, days)
    for row in rows:
        counters[row.production_day] += 1
        row.production_sequence = counters[row.production_day]


def sync_parsed_rows_to_db(
    cursor,
    rows: list[ParsedWeldingRow],
    *,
    dry_run: bool = False,
) -> ImportSyncResult:
    result = ImportSyncResult(parsed=len(rows))
    if not rows:
        return result

    keys = [r.external_sync_key for r in rows]
    existing_keys = _load_existing_sync_keys(cursor, keys)
    new_rows = [r for r in rows if r.external_sync_key not in existing_keys]

    biz_fps = _load_existing_business_fingerprints(cursor, {r.production_day for r in new_rows})
    deduped: list[ParsedWeldingRow] = []
    for row in new_rows:
        fp = business_fingerprint(
            row.production_day,
            row.product_cd,
            row.welding_machine,
            row.operator_user_id,
            row.actual_qty,
            row.defect_qty,
        )
        if fp in biz_fps:
            continue
        deduped.append(row)
        biz_fps.add(fp)
    result.skipped_duplicate = len(rows) - len(deduped)
    new_rows = deduped

    for r in rows:
        if r.operator_user_id is None:
            result.unmapped_operators[r.operator_name] = result.unmapped_operators.get(r.operator_name, 0) + 1

    if not new_rows or dry_run:
        result.inserted = len(new_rows) if dry_run else 0
        return result

    assign_production_sequences(new_rows, cursor)
    has_data_source_col = _table_has_column(cursor, "welding_management", "data_source")
    has_sync_col = _table_has_column(cursor, "welding_management", "external_sync_key")
    has_machine_col = _table_has_column(cursor, "welding_management", "welding_machine")

    insert_cols = [
        "production_month",
        "production_day",
        "production_sequence",
        "product_cd",
        "product_name",
    ]
    if has_machine_col:
        insert_cols.append("welding_machine")
    insert_cols.extend(
        [
            "actual_production_quantity",
            "defect_qty",
            "mes_defect_by_item",
            "production_completed_check",
            "mes_production_started_at",
            "mes_production_ended_at",
            "mes_net_production_sec",
            "mes_paused_accum_sec",
        ]
    )
    if _table_has_column(cursor, "welding_management", "mes_shift_sec"):
        insert_cols.extend(["mes_shift_sec", "mes_break_sec", "mes_stop_sec"])
    insert_cols.extend(
        [
            "mes_production_is_paused",
            "mes_operator_user_id",
            "remarks",
        ]
    )
    if has_sync_col:
        insert_cols.append("external_sync_key")
    if has_data_source_col:
        insert_cols.append("data_source")
    placeholders = ", ".join(["%s"] * len(insert_cols))
    sql = f"INSERT INTO welding_management ({', '.join(insert_cols)}) VALUES ({placeholders})"

    batch_params = []
    for row in new_rows:
        values: list[Any] = [
            row.production_month,
            row.production_day,
            row.production_sequence,
            row.product_cd,
            row.product_name,
        ]
        if has_machine_col:
            values.append(row.welding_machine or None)
        values.extend(
            [
                row.actual_qty,
                row.defect_qty,
                None,
                1,
                row.mes_production_started_at,
                row.mes_production_ended_at,
                row.mes_net_production_sec,
                row.mes_paused_accum_sec,
            ]
        )
        if _table_has_column(cursor, "welding_management", "mes_shift_sec"):
            values.extend([row.mes_shift_sec, row.mes_break_sec, row.mes_stop_sec])
        values.extend(
            [
                0,
                row.operator_user_id,
                row.remarks,
            ]
        )
        if has_sync_col:
            values.append(row.external_sync_key)
        if has_data_source_col:
            values.append(row.data_source)
        batch_params.append(tuple(values))

    cursor.executemany(sql, batch_params)
    result.inserted = len(new_rows)
    return result


def delete_rows_by_remarks_prefix(cursor, prefix: str, dry_run: bool = False) -> int:
    pattern = f"{prefix}%"
    cursor.execute(
        "SELECT COUNT(1) FROM welding_management WHERE remarks LIKE %s",
        (pattern,),
    )
    count = int(cursor.fetchone()[0])
    if count and not dry_run:
        cursor.execute("DELETE FROM welding_management WHERE remarks LIKE %s", (pattern,))
    return count


def sync_welding_source_file(
    source_path: str,
    connection_factory,
    *,
    dry_run: bool = False,
    skip_unmapped_operator: bool = False,
    worker_overrides: dict[str, int] | None = None,
    remarks_prefix: str = REMARKS_PREFIX_EXCEL,
) -> ImportSyncResult:
    """CSV/Excel 1 ファイル（入力表）を welding_management に增量同期（内容ハッシュで重複排除）"""
    conn = connection_factory()
    conn.autocommit = False
    cursor = conn.cursor()
    try:
        user_rows = load_user_rows(cursor)
        rows, parse_errors = parse_rows_from_source(
            source_path,
            user_rows=user_rows,
            worker_overrides=worker_overrides,
            skip_unmapped_operator=skip_unmapped_operator,
            remarks_prefix=remarks_prefix,
        )
        result = sync_parsed_rows_to_db(cursor, rows, dry_run=dry_run)
        result.errors.extend(parse_errors)
        if not dry_run:
            conn.commit()
        logger.info(
            "welding_management 同期完了: parsed=%s inserted=%s dup_skip=%s path=%s",
            result.parsed,
            result.inserted,
            result.skipped_duplicate,
            source_path,
        )
        return result
    except Exception:
        conn.rollback()
        raise
    finally:
        cursor.close()
        conn.close()
