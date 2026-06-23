"""レポート生成の共通基盤（データ構造・相対日付解決・Excel ユーティリティ）"""
from __future__ import annotations

import io
from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from datetime import date, timedelta
from typing import Any

from sqlalchemy.ext.asyncio import AsyncSession

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PDF_MIME = "application/pdf"


@dataclass(frozen=True)
class ReportAttachment:
    filename: str
    content: bytes
    mime_type: str


@dataclass
class GeneratedReport:
    """生成済みレポート（添付 + メール/LINE 用要約）。"""

    period_label: str
    record_count: int
    summary_html: str
    summary_text: str
    attachments: list[ReportAttachment] = field(default_factory=list)
    extra_variables: dict[str, Any] = field(default_factory=dict)


@dataclass(frozen=True)
class ResolvedRange:
    start: date
    end: date
    label: str


def _last_week_range(today: date) -> tuple[date, date]:
    """前週（月曜〜日曜）。"""
    this_monday = today - timedelta(days=today.weekday())
    last_monday = this_monday - timedelta(days=7)
    return last_monday, last_monday + timedelta(days=6)


def _this_week_range(today: date) -> tuple[date, date]:
    monday = today - timedelta(days=today.weekday())
    return monday, monday + timedelta(days=6)


def _last_month_range(today: date) -> tuple[date, date]:
    first_this = today.replace(day=1)
    last_prev_end = first_this - timedelta(days=1)
    return last_prev_end.replace(day=1), last_prev_end


def _this_month_range(today: date) -> tuple[date, date]:
    first = today.replace(day=1)
    if first.month == 12:
        next_first = first.replace(year=first.year + 1, month=1)
    else:
        next_first = first.replace(month=first.month + 1)
    return first, next_first - timedelta(days=1)


def resolve_date_range(parameters: dict[str, Any], *, run_date: date) -> ResolvedRange:
    """parameters の date_range（プリセット or custom）を具体的な期間に解決する。"""
    raw = parameters.get("date_range") or "yesterday"

    if raw == "custom":
        start = _parse_date(parameters.get("start_date")) or (run_date - timedelta(days=1))
        end = _parse_date(parameters.get("end_date")) or start
        label = f"{start.isoformat()} 〜 {end.isoformat()}"
        return ResolvedRange(start=start, end=end, label=label)

    if raw == "today":
        return ResolvedRange(start=run_date, end=run_date, label=run_date.isoformat())
    if raw == "yesterday":
        d = run_date - timedelta(days=1)
        return ResolvedRange(start=d, end=d, label=d.isoformat())
    if raw == "last_week":
        s, e = _last_week_range(run_date)
        return ResolvedRange(start=s, end=e, label=f"{s.isoformat()} 〜 {e.isoformat()}")
    if raw == "this_week":
        s, e = _this_week_range(run_date)
        return ResolvedRange(start=s, end=e, label=f"{s.isoformat()} 〜 {e.isoformat()}")
    if raw == "last_month":
        s, e = _last_month_range(run_date)
        return ResolvedRange(start=s, end=e, label=f"{s.strftime('%Y-%m')}")
    if raw == "this_month":
        s, e = _this_month_range(run_date)
        return ResolvedRange(start=s, end=e, label=f"{s.strftime('%Y-%m')}")

    # 未知のトークンは前日扱い
    d = run_date - timedelta(days=1)
    return ResolvedRange(start=d, end=d, label=d.isoformat())


def resolve_month(parameters: dict[str, Any], *, run_date: date) -> date:
    """parameters の month（last_month|this_month|YYYY-MM）を月初日に解決する。"""
    raw = parameters.get("month") or "last_month"
    if raw == "this_month":
        return run_date.replace(day=1)
    if raw == "last_month":
        s, _ = _last_month_range(run_date)
        return s
    parsed = _parse_date(raw if len(str(raw)) > 7 else f"{raw}-01")
    if parsed:
        return parsed.replace(day=1)
    s, _ = _last_month_range(run_date)
    return s


def _parse_date(value: Any) -> date | None:
    if not value:
        return None
    try:
        parts = str(value).strip()[:10].split("-")
        if len(parts) != 3:
            return None
        return date(int(parts[0]), int(parts[1]), int(parts[2]))
    except (ValueError, IndexError):
        return None


def build_xlsx(sheets: list[tuple[str, list[str], list[list[Any]]]]) -> bytes:
    """(シート名, ヘッダー, 行) のリストから xlsx バイト列を生成する。"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E8EEF7")

    for name, headers, rows in sheets:
        ws = wb.create_sheet(title=name[:31] or "Sheet")
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for row in rows:
            ws.append(row)
        for idx, header in enumerate(headers, start=1):
            max_len = len(str(header))
            for row in rows:
                if idx - 1 < len(row):
                    max_len = max(max_len, len(str(row[idx - 1])))
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = min(max_len + 4, 50)

    if not wb.sheetnames:
        wb.create_sheet(title="Sheet")
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


_JP_FONT = "HeiseiKakuGo-W5"
_jp_font_registered = False


def _ensure_jp_font() -> str:
    """reportlab 内蔵の日本語 CID フォントを登録する（フォントファイル不要）。"""
    global _jp_font_registered
    if not _jp_font_registered:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont

        pdfmetrics.registerFont(UnicodeCIDFont(_JP_FONT))
        _jp_font_registered = True
    return _JP_FONT


def build_pdf(
    title: str,
    period_label: str,
    sections: list[tuple[str, list[str], list[list[Any]]]],
) -> bytes:
    """(見出し, ヘッダー, 行) のセクションから日本語対応 PDF を生成する。"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    font = _ensure_jp_font()
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("jp_title", parent=styles["Title"], fontName=font, fontSize=15)
    head_style = ParagraphStyle("jp_head", parent=styles["Heading2"], fontName=font, fontSize=11)
    body_style = ParagraphStyle("jp_body", parent=styles["Normal"], fontName=font, fontSize=9)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=15 * mm, bottomMargin=15 * mm)
    flow: list[Any] = [Paragraph(title, title_style), Paragraph(f"対象期間: {period_label}", body_style), Spacer(1, 8)]

    for heading, headers, rows in sections:
        flow.append(Paragraph(heading, head_style))
        table_data = [headers] + [[str(c) for c in row] for row in rows]
        table = Table(table_data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, -1), font),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8EEF7")),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )
        flow.append(table)
        flow.append(Spacer(1, 12))

    doc.build(flow)
    return buffer.getvalue()


class ReportGenerator(ABC):
    """各レポートタイプが実装する生成器。"""

    report_code: str

    @abstractmethod
    async def generate(
        self,
        db: AsyncSession,
        *,
        parameters: dict[str, Any],
        fmt: str,
        run_date: date,
    ) -> GeneratedReport:
        ...

    def reference_key(self, *, parameters: dict[str, Any], run_date: date) -> str:
        """重複送信防止用キー。既定は対象期間ベース。"""
        rng = resolve_date_range(parameters, run_date=run_date)
        return f"{self.report_code}:{rng.start.isoformat()}:{rng.end.isoformat()}"
