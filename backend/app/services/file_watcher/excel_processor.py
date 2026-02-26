# coding: utf-8
"""生産計画 Excel 監視処理（excelFileMonitorService.js と同等ロジック）：加工・溶接 .xlsm → production_plan_*"""
import os
import re
import time
import logging
import warnings
from datetime import datetime, timedelta, date

from app.services.file_watcher.sync_services import get_db_connection

logger = logging.getLogger(__name__)

EXCEL_FILES = []
for i in range(1, 13):
    EXCEL_FILES.append(f"加工計画({i}月).xlsm")
    EXCEL_FILES.append(f"溶接計画({i}月).xlsm")
EXCEL_FILES_SET = set(EXCEL_FILES)

# 計画更新：日期列名优先级（与 JS isDateInRange 一致）
DATE_FIELDS = ["生産日", "計画日", "予定日", "日付", "日期", "date", "Date"]

BATCH_INSERT_SIZE = 500


def _normalize_excel_filename(name):
    if not name:
        return ""
    return name.replace("\uFF08", "(").replace("\uFF09", ")")


def is_excel_target_file(filename):
    return _normalize_excel_filename(filename) in EXCEL_FILES_SET


def parse_excel_date(value):
    """Excel 日期：datetime/数字(序列日)/字符串"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date() if hasattr(value, "date") else value
    if isinstance(value, (int, float)):
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
    if isinstance(value, str):
        s = value.strip()[:10]
        for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y%m%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def parse_number(value):
    """与 JS parseNumber 一致：空→None，否则数值"""
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return value if (value == value) else None  # NaN check
    try:
        n = float(value)
        return n if (n == n) else None
    except (ValueError, TypeError):
        return None


def clean_cell_value(value):
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip()
    return value


def extract_month_from_filename(filename):
    name = _normalize_excel_filename(filename or "")
    m = re.search(r"(?:加工計画|溶接計画)\((\d+)月\)", name)
    if not m:
        return None
    try:
        month = int(m.group(1))
        return month if 1 <= month <= 12 else None
    except (ValueError, IndexError):
        return None


def get_date_range_for_month(month):
    """当年该月1日～翌年该月最后一天（与 JS getDateRangeForMonth 一致）"""
    if month is None or not (1 <= month <= 12):
        return None, None
    today = date.today()
    start = date(today.year, month, 1)
    if month == 12:
        end = date(today.year + 1, 12, 31)
    else:
        end = date(today.year + 1, month + 1, 1) - timedelta(days=1)
    return start, end


def _executemany_batch(cursor, sql, rows, batch_size=BATCH_INSERT_SIZE):
    for i in range(0, len(rows), batch_size):
        cursor.executemany(sql, rows[i : i + batch_size])


class ExcelProcessor:
    def is_target_file(self, filename):
        return is_excel_target_file(filename)

    def process_file(self, filepath):
        try:
            import openpyxl  # type: ignore[import-untyped]
        except ImportError:
            logger.error("openpyxl が未インストールです。pip install openpyxl を実行してください")
            raise
        filename = os.path.basename(filepath)
        logger.info("開始処理 Excel: %s", filename)
        conn = None
        wb = None
        try:
            # openpyxl の「印刷範囲＝定義名」警告を抑制（データ読取には影響なし）
            with warnings.catch_warnings():
                warnings.filterwarnings("ignore", message="Print area cannot be set to Defined name")
                wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
            max_retries = 3
            for attempt in range(max_retries):
                conn = get_db_connection()
                conn.autocommit = False
                cursor = conn.cursor()
                try:
                    cursor.execute("DELETE FROM production_plan_updates WHERE file_name = %s", (filename,))
                    cursor.execute("DELETE FROM production_plan_schedules WHERE file_name = %s", (filename,))
                    cursor.execute("DELETE FROM production_plan_rate WHERE file_name = %s", (filename,))

                    if "計画更新" in wb.sheetnames:
                        self._process_plan_update(wb["計画更新"], filename, cursor)
                    else:
                        logger.warning("%s に '計画更新' シートがありません", filename)

                    if "加工状況" in wb.sheetnames:
                        self._process_processing_status(wb["加工状況"], filename, cursor)
                    if "溶接状況" in wb.sheetnames:
                        self._process_welding_status(wb["溶接状況"], filename, cursor)
                    if "操業度" in wb.sheetnames:
                        self._process_operation_rate(wb["操業度"], filename, cursor)

                    conn.commit()
                    logger.info("Excel 処理完了: %s", filename)
                    break
                except Exception as e:
                    if conn:
                        conn.rollback()
                    is_deadlock = getattr(e, "errno", None) == 1213 or "Deadlock" in str(e) or "1213" in str(e)
                    if is_deadlock and attempt < max_retries - 1:
                        logger.warning("Excel 死锁重试 %s/%s: %s", attempt + 1, max_retries, filename)
                        time.sleep(0.5 * (attempt + 1))
                    else:
                        raise
                finally:
                    if conn:
                        try:
                            conn.close()
                        except Exception:
                            pass
                    conn = None
        except Exception as e:
            logger.error("Excel 処理失敗 %s: %s", filename, e, exc_info=True)
        finally:
            if conn is not None:
                try:
                    conn.close()
                except Exception:
                    pass
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass

    def _process_plan_update(self, sheet, filename, cursor):
        """計画更新：ヘッダー 生産日/生産数/ライン/生産準/品名；日付範囲はファイル名の月；machine_cd/product_cd は一括取得で高速化"""
        headers = {}
        for cell in sheet[1]:
            if cell.value:
                headers[str(cell.value).strip()] = cell.column - 1

        def get_val(row, name):
            idx = headers.get(name)
            return row[idx] if idx is not None and idx < len(row) else None

        process_name = "溶接" if "溶接" in filename else "成型"
        month = extract_month_from_filename(filename)
        date_start, date_end = get_date_range_for_month(month) if month else (None, None)

        # 第一遍：解析行、日期范围过滤、收集唯一 machine_name / product_name
        candidate_rows = []
        machine_names = set()
        product_names = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row = list(row) if row else []
            plan_date = None
            for field in DATE_FIELDS:
                v = get_val(row, field)
                plan_date = parse_excel_date(v)
                if plan_date is not None:
                    break
            quantity = get_val(row, "生産数")
            machine_name = clean_cell_value(get_val(row, "ライン"))
            operator = clean_cell_value(get_val(row, "生産準"))
            product_name = clean_cell_value(get_val(row, "品名"))
            if not (plan_date and quantity is not None and machine_name and product_name):
                continue
            qty = parse_number(quantity)
            if qty is not None and qty <= 0:
                continue
            if date_start is not None and date_end is not None and not (date_start <= plan_date <= date_end):
                continue
            if "加工" in (machine_name or ""):
                machine_name = (machine_name or "").replace("加工", "成型")
            machine_names.add(machine_name)
            product_names.add(product_name)
            candidate_rows.append((plan_date, qty or quantity, machine_name, operator, product_name))

        if not candidate_rows:
            return

        # 批量查 machine_cd / product_cd（与 JS 一致，减少 N 次查询）
        machine_cd_map = {}
        if machine_names:
            placeholders = ",".join(["%s"] * len(machine_names))
            cursor.execute(
                f"SELECT machine_name, machine_cd FROM machines WHERE machine_name IN ({placeholders})",
                list(machine_names),
            )
            for r in cursor.fetchall():
                machine_cd_map[r[0]] = r[1]
        product_cd_map = {}
        if product_names:
            placeholders = ",".join(["%s"] * len(product_names))
            cursor.execute(
                f"SELECT product_name, product_cd FROM products WHERE product_name IN ({placeholders}) AND RIGHT(product_cd, 1) = '1'",
                list(product_names),
            )
            for r in cursor.fetchall():
                product_cd_map[r[0]] = r[1]

        insert_sql = """
            INSERT INTO production_plan_updates
            (file_name, processed_at, plan_date, quantity, machine_name, machine_cd, process_name, operator, product_name, product_cd)
            VALUES (%s, NOW(), %s, %s, %s, %s, %s, %s, %s, %s)
        """
        rows_to_insert = []
        for plan_date, qty, machine_name, operator, product_name in candidate_rows:
            machine_cd = machine_cd_map.get(machine_name)
            product_cd = product_cd_map.get(product_name)
            rows_to_insert.append((
                filename, plan_date, qty, machine_name, machine_cd, process_name, operator, product_name, product_cd,
            ))
        _executemany_batch(cursor, insert_sql, rows_to_insert)
        logger.info("計画更新: %s 行を挿入", len(rows_to_insert))

    def _process_processing_status(self, sheet, filename, cursor):
        """加工状況：A=生産順序,B=設備名,C=生産品種,D=計画数,E/F=開始/終了日,G=実績,H=差異,I=達成率,J=総時間,K=操業度差異,L=材料ロット,M=材料名（JS と同一列）"""
        insert_sql = """
            INSERT INTO production_plan_schedules
            (file_name, processed_at, machine_name, product_name, production_order, planned_quantity,
             production_start_date, production_end_date, actual_production, variance,
             achievement_rate, total_production_time, operation_variance, material_lot_count, material_name)
            VALUES (%s, NOW(), %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        rows_to_insert = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row = list(row) if row else []
            if len(row) < 3:
                continue
            # JS: 有効数据需 machine_name || product_name || planned_quantity || actual_production；且 map 时无 product_name 则 return null
            if not (row[1] or row[2]):
                continue
            if not row[2]:
                continue
            machine_name = str(row[1] or "").replace("加工", "成型")
            product_name = str(row[2]).strip() if row[2] else None
            planned_qty = parse_number(row[3]) if len(row) > 3 else None
            start_date = parse_excel_date(row[4]) if len(row) > 4 else None
            end_date = parse_excel_date(row[5]) if len(row) > 5 else None
            actual = parse_number(row[6]) if len(row) > 6 else None
            variance = parse_number(row[7]) if len(row) > 7 else None
            achievement = row[8] if len(row) > 8 else None
            if achievement is not None and isinstance(achievement, str) and "%" in achievement:
                achievement = parse_number(achievement.replace("%", "").strip())
            else:
                achievement = parse_number(achievement)
            total_time = parse_number(row[9]) if len(row) > 9 else None
            op_var = parse_number(row[10]) if len(row) > 10 else None
            mat_count = parse_number(row[11]) if len(row) > 11 else None
            mat_name = str(row[12]).strip() if len(row) > 12 and row[12] else None
            rows_to_insert.append((
                filename, machine_name, product_name,
                parse_number(row[0]) if len(row) > 0 else None,
                planned_qty, start_date, end_date, actual, variance,
                achievement, total_time, op_var, mat_count, mat_name,
            ))
        if rows_to_insert:
            _executemany_batch(cursor, insert_sql, rows_to_insert)
            logger.info("加工状況: %s 行を挿入", len(rows_to_insert))

    def _process_welding_status(self, sheet, filename, cursor):
        self._process_processing_status(sheet, filename, cursor)

    def _process_operation_rate(self, sheet, filename, cursor):
        """操業度：A=設備CD,B=設備名,C=操業度差異；A が空のとき B から machine_cd を参照（JS と同様）"""
        rows_raw = []
        machine_names_for_lookup = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row = list(row) if row else []
            if len(row) < 2:
                continue
            machine_cd = str(row[0]).strip() if len(row) > 0 and row[0] else None
            machine_name = str(row[1]).replace("加工", "成型") if len(row) > 1 and row[1] else None
            op_var = parse_number(row[2]) if len(row) > 2 else None
            if not (machine_cd or machine_name):
                continue
            rows_raw.append((machine_cd, machine_name, op_var))
            if not machine_cd and machine_name:
                machine_names_for_lookup.add(machine_name)

        machine_cd_by_name = {}
        if machine_names_for_lookup:
            placeholders = ",".join(["%s"] * len(machine_names_for_lookup))
            cursor.execute(
                f"SELECT machine_name, machine_cd FROM machines WHERE machine_name IN ({placeholders})",
                list(machine_names_for_lookup),
            )
            for r in cursor.fetchall():
                machine_cd_by_name[r[0]] = r[1]

        insert_sql = """
            INSERT INTO production_plan_rate (file_name, processed_at, machine_cd, machine_name, operation_variance)
            VALUES (%s, NOW(), %s, %s, %s)
        """
        rows_to_insert = []
        for machine_cd, machine_name, op_var in rows_raw:
            if not machine_cd and machine_name:
                machine_cd = machine_cd_by_name.get(machine_name)
            rows_to_insert.append((filename, machine_cd, machine_name, op_var))
        if rows_to_insert:
            _executemany_batch(cursor, insert_sql, rows_to_insert)
            logger.info("操業度: %s 行を挿入", len(rows_to_insert))
