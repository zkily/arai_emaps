# coding: utf-8
"""生産管理指標(YYYY年度-切断) CSV/Excel → cutting_production_indicator 取込（CLI・ファイル監視共用）"""
from __future__ import annotations

import csv
import hashlib
import logging
import os
import re
import warnings
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterator

logger = logging.getLogger(__name__)

EXTERNAL_SYNC_KEY_PREFIX = "cut_excel:"
REMARKS_PREFIX_EXCEL = "EXCEL_SYNC"
REMARKS_PREFIX_CSV = "CSV_IMPORT"

DATA_SOURCE_EXCEL = "excel"
DATA_SOURCE_CSV = "csv"

# 列位置（CSV ヘッダ重複「差異」対策で位置ベース）
COLUMN_FIELD_NAMES = [
    "product_cd",
    "production_day",
    "production_line",
    "product_name",
    "planned_quantity",
    "actual_quantity",
    "quantity_variance",
    "shift_hours",
    "break_hours",
    "setup_hours",
    "repair_hours",
    "saw_blade_exchange_hours",
    "loss_stop_day_hours",
    "loss_stop_night_hours",
    "invisible_loss_stop_hours",
    "planned_stop_hours",
    "available_work_hours",
    "work_hours",
    "utilization_rate",
    "work_rate",
    "total_production_qty",
    "efficiency_rate",
    "equipment_record_time_1",
    "equipment_record_time_2",
    "metric_60",
    "spare_col_1",
    "spare_col_2",
    "actual_record_cutting",
    "cutting_record_variance",
    "actual_record_chamfering",
    "chamfering_record_variance",
    "chamfering_setup_hours",
    "spare_col_3",
    "forming_process_date",
]

CUTTING_SHEET_HEADER_MARKERS = ("日付", "ライン", "品名", "生産数")
FISCAL_YEAR_RE = re.compile(r"(\d{4})年度")
PRODUCT_CD_RE = re.compile(r"^\d{4,6}$")

INT_FIELDS = frozenset(
    {
        "planned_quantity",
        "actual_quantity",
        "quantity_variance",
        "total_production_qty",
    }
)
DECIMAL_FIELDS = frozenset(
    {
        "shift_hours",
        "break_hours",
        "setup_hours",
        "repair_hours",
        "saw_blade_exchange_hours",
        "loss_stop_day_hours",
        "loss_stop_night_hours",
        "invisible_loss_stop_hours",
        "planned_stop_hours",
        "available_work_hours",
        "work_hours",
        "utilization_rate",
        "work_rate",
        "efficiency_rate",
        "equipment_record_time_1",
        "equipment_record_time_2",
        "actual_record_cutting",
        "cutting_record_variance",
        "actual_record_chamfering",
        "chamfering_record_variance",
        "chamfering_setup_hours",
    }
)
DATE_FIELDS = frozenset({"production_day", "forming_process_date"})


def _norm_header(s: str) -> str:
    return (s or "").strip().replace("\u3000", " ").strip()


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y/%m/%d") if value.hour == 0 and value.minute == 0 else value.isoformat()
    if isinstance(value, date) and not isinstance(value, datetime):
        return value.strftime("%Y/%m/%d")
    return str(value).strip()


def _parse_int(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value != value:
            return None
        return int(value)
    s = _cell_str(value).replace(",", "").replace("，", "")
    if not s or s in ("-", "—", "#N/A"):
        return None
    try:
        return int(float(s))
    except (TypeError, ValueError):
        return None


def _parse_decimal(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        n = float(value)
        return n if n == n else None
    s = _cell_str(value).replace(",", "").replace("，", "").replace("%", "").strip()
    if not s or s in ("-", "—", "#N/A"):
        return None
    try:
        n = float(s)
        return n if n == n else None
    except (TypeError, ValueError):
        return None


def _parse_date(value: Any, fiscal_year: int | None) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = _cell_str(value)
    if not s or s in ("#N/A", "-", "—"):
        return None
    for fmt in ("%Y/%m/%d", "%Y-%m-%d"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            pass
    m = re.match(r"^(\d{4})[/-](\d{1,2})[/-](\d{1,2})$", s)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    m = re.match(r"^(\d{1,2})[/-](\d{1,2})$", s)
    if m and fiscal_year:
        try:
            return date(fiscal_year, int(m.group(1)), int(m.group(2)))
        except ValueError:
            return None
    return None


def fiscal_year_from_source(source: str) -> int | None:
    m = FISCAL_YEAR_RE.search((source or "").replace("\uFF08", "(").replace("\uFF09", ")"))
    if not m:
        return None
    return int(m.group(1))


def row_values_to_fields(values: tuple[Any, ...], fiscal_year: int | None) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for i, field_name in enumerate(COLUMN_FIELD_NAMES):
        if field_name.startswith("spare_col"):
            continue
        raw = values[i] if i < len(values) else None
        if field_name in DATE_FIELDS:
            out[field_name] = _parse_date(raw, fiscal_year)
        elif field_name in INT_FIELDS:
            out[field_name] = _parse_int(raw)
        elif field_name in DECIMAL_FIELDS:
            out[field_name] = _parse_decimal(raw)
        elif field_name == "metric_60":
            s = _cell_str(raw)
            out[field_name] = s if s else None
        else:
            s = _cell_str(raw)
            out[field_name] = s if s else None
    return out


def is_data_row(fields: dict[str, Any]) -> bool:
    if fields.get("production_day") is None:
        return False
    line = (fields.get("production_line") or "").strip()
    if not line:
        return False
    cd = (fields.get("product_cd") or "").strip()
    if cd.startswith("#") or cd == "集計":
        return False
    name = (fields.get("product_name") or "").strip()
    if cd and not PRODUCT_CD_RE.match(cd):
        if not name:
            return False
    planned = fields.get("planned_quantity") or 0
    actual = fields.get("actual_quantity") or 0
    work_h = fields.get("work_hours")
    if not name and not cd and planned == 0 and actual == 0 and (work_h is None or work_h <= 0):
        return False
    return True


@dataclass
class ParsedCuttingIndicatorRow:
    source_line: int
    fiscal_year: int | None
    production_month: date | None
    production_day: date
    product_cd: str | None
    production_line: str | None
    product_name: str | None
    planned_quantity: int | None
    actual_quantity: int | None
    quantity_variance: int | None
    shift_hours: float | None
    break_hours: float | None
    setup_hours: float | None
    repair_hours: float | None
    saw_blade_exchange_hours: float | None
    loss_stop_day_hours: float | None
    loss_stop_night_hours: float | None
    invisible_loss_stop_hours: float | None
    planned_stop_hours: float | None
    available_work_hours: float | None
    work_hours: float | None
    utilization_rate: float | None
    work_rate: float | None
    total_production_qty: int | None
    efficiency_rate: float | None
    equipment_record_time_1: float | None
    equipment_record_time_2: float | None
    metric_60: str | None
    actual_record_cutting: float | None
    cutting_record_variance: float | None
    actual_record_chamfering: float | None
    chamfering_record_variance: float | None
    chamfering_setup_hours: float | None
    forming_process_date: date | None
    external_sync_key: str
    data_source: str
    remarks: str
    source_file: str


@dataclass
class ImportSyncResult:
    parsed: int = 0
    inserted: int = 0
    deleted: int = 0
    skipped_invalid: int = 0
    errors: list[str] = field(default_factory=list)


def make_external_sync_key(source_file: str, source_line: int, fields: dict[str, Any]) -> str:
    parts = [
        source_file,
        str(source_line),
        (fields.get("production_day") or "").isoformat() if fields.get("production_day") else "",
        (fields.get("production_line") or "").strip(),
        (fields.get("product_cd") or "").strip(),
        (fields.get("product_name") or "").strip(),
        str(fields.get("planned_quantity") or ""),
        str(fields.get("actual_quantity") or ""),
        str(fields.get("shift_hours") or ""),
        str(fields.get("work_hours") or ""),
    ]
    digest = hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()[:40]
    return f"{EXTERNAL_SYNC_KEY_PREFIX}{digest}"


def parse_logical_row(
    fields: dict[str, Any],
    *,
    source_line: int,
    source_label: str,
    remarks_prefix: str,
    fiscal_year: int | None,
) -> ParsedCuttingIndicatorRow | None:
    if not is_data_row(fields):
        return None
    prod_day = fields["production_day"]
    sync_key = make_external_sync_key(source_label, source_line, fields)
    return ParsedCuttingIndicatorRow(
        source_line=source_line,
        fiscal_year=fiscal_year,
        production_month=prod_day.replace(day=1) if prod_day else None,
        production_day=prod_day,
        product_cd=(fields.get("product_cd") or "").strip() or None,
        production_line=(fields.get("production_line") or "").strip() or None,
        product_name=(fields.get("product_name") or "").strip() or None,
        planned_quantity=fields.get("planned_quantity"),
        actual_quantity=fields.get("actual_quantity"),
        quantity_variance=fields.get("quantity_variance"),
        shift_hours=fields.get("shift_hours"),
        break_hours=fields.get("break_hours"),
        setup_hours=fields.get("setup_hours"),
        repair_hours=fields.get("repair_hours"),
        saw_blade_exchange_hours=fields.get("saw_blade_exchange_hours"),
        loss_stop_day_hours=fields.get("loss_stop_day_hours"),
        loss_stop_night_hours=fields.get("loss_stop_night_hours"),
        invisible_loss_stop_hours=fields.get("invisible_loss_stop_hours"),
        planned_stop_hours=fields.get("planned_stop_hours"),
        available_work_hours=fields.get("available_work_hours"),
        work_hours=fields.get("work_hours"),
        utilization_rate=fields.get("utilization_rate"),
        work_rate=fields.get("work_rate"),
        total_production_qty=fields.get("total_production_qty"),
        efficiency_rate=fields.get("efficiency_rate"),
        equipment_record_time_1=fields.get("equipment_record_time_1"),
        equipment_record_time_2=fields.get("equipment_record_time_2"),
        metric_60=fields.get("metric_60"),
        actual_record_cutting=fields.get("actual_record_cutting"),
        cutting_record_variance=fields.get("cutting_record_variance"),
        actual_record_chamfering=fields.get("actual_record_chamfering"),
        chamfering_record_variance=fields.get("chamfering_record_variance"),
        chamfering_setup_hours=fields.get("chamfering_setup_hours"),
        forming_process_date=fields.get("forming_process_date"),
        external_sync_key=sync_key,
        data_source=DATA_SOURCE_CSV if remarks_prefix == REMARKS_PREFIX_CSV else DATA_SOURCE_EXCEL,
        remarks=f"{remarks_prefix}:{source_label}:L{source_line}",
        source_file=source_label,
    )


def iter_csv_rows(csv_path: Path) -> Iterator[tuple[int, dict[str, Any]]]:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        header = next(reader, None)
        if not header:
            return
        fiscal_year = fiscal_year_from_source(csv_path.name)
        for line_no, values in enumerate(reader, start=2):
            if not values:
                continue
            yield line_no, row_values_to_fields(tuple(values), fiscal_year)


def _sheet_has_cutting_headers(sheet) -> bool:
    rows_iter = sheet.iter_rows(min_row=1, max_row=1, values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        return False
    headers = {_norm_header(_cell_str(h)) for h in header_row if _cell_str(h)}
    return all(marker in headers for marker in CUTTING_SHEET_HEADER_MARKERS)


def resolve_cutting_data_sheet(wb, *, filepath: str = ""):
    for name in wb.sheetnames:
        sheet = wb[name]
        if _sheet_has_cutting_headers(sheet):
            return name, sheet
    active = wb.active
    if active is not None and _sheet_has_cutting_headers(active):
        return active.title, active
    available = ", ".join(wb.sheetnames)
    location = f"（{filepath}）" if filepath else ""
    raise ValueError(
        f"切断生産管理指標のデータシート（日付/ライン/品名/生産数 ヘッダ）が見つかりません"
        f"{location}。利用可能: {available}"
    )


def iter_cutting_excel_rows(filepath: str) -> Iterator[tuple[int, dict[str, Any]]]:
    import openpyxl

    fiscal_year = fiscal_year_from_source(os.path.basename(filepath))
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="Print area cannot be set to Defined name")
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    try:
        _resolved_name, sheet = resolve_cutting_data_sheet(wb, filepath=filepath)
        if sheet is None or (sheet.max_row or 0) <= 1:
            return
        rows_iter = sheet.iter_rows(min_row=2, values_only=True)
        for line_no, values in enumerate(rows_iter, start=2):
            if not values:
                continue
            yield line_no, row_values_to_fields(values, fiscal_year)
    finally:
        try:
            wb.close()
        except Exception:
            pass


def parse_rows_from_source(
    source_path: str,
    *,
    remarks_prefix: str = REMARKS_PREFIX_EXCEL,
) -> tuple[list[ParsedCuttingIndicatorRow], list[str], int]:
    path = Path(source_path)
    source_label = path.name
    fiscal_year = fiscal_year_from_source(source_label)
    parsed: list[ParsedCuttingIndicatorRow] = []
    errors: list[str] = []
    skipped = 0

    suffix = path.suffix.lower()
    if suffix == ".csv":
        row_iter = iter_csv_rows(path)
    elif suffix in (".xlsx", ".xlsm"):
        row_iter = iter_cutting_excel_rows(str(path))
    else:
        raise ValueError(f"未対応の拡張子: {suffix}")

    for line_no, fields in row_iter:
        if not is_data_row(fields):
            skipped += 1
            continue
        row = parse_logical_row(
            fields,
            source_line=line_no,
            source_label=source_label,
            remarks_prefix=remarks_prefix,
            fiscal_year=fiscal_year,
        )
        if row is None:
            skipped += 1
            errors.append(f"行 {line_no}: 解析不可")
            continue
        parsed.append(row)
    return parsed, errors, skipped


INSERT_COLUMNS = [
    "fiscal_year",
    "production_month",
    "production_day",
    "source_line",
    "source_file",
    "product_cd",
    "production_line",
    "product_name",
    "planned_quantity",
    "actual_quantity",
    "quantity_variance",
    "shift_hours",
    "break_hours",
    "setup_hours",
    "repair_hours",
    "saw_blade_exchange_hours",
    "loss_stop_day_hours",
    "loss_stop_night_hours",
    "invisible_loss_stop_hours",
    "planned_stop_hours",
    "available_work_hours",
    "work_hours",
    "utilization_rate",
    "work_rate",
    "total_production_qty",
    "efficiency_rate",
    "equipment_record_time_1",
    "equipment_record_time_2",
    "metric_60",
    "actual_record_cutting",
    "cutting_record_variance",
    "actual_record_chamfering",
    "chamfering_record_variance",
    "chamfering_setup_hours",
    "forming_process_date",
    "data_source",
    "external_sync_key",
    "remarks",
]


def _row_to_insert_tuple(row: ParsedCuttingIndicatorRow) -> tuple:
    return (
        row.fiscal_year,
        row.production_month,
        row.production_day,
        row.source_line,
        row.source_file,
        row.product_cd,
        row.production_line,
        row.product_name,
        row.planned_quantity,
        row.actual_quantity,
        row.quantity_variance,
        row.shift_hours,
        row.break_hours,
        row.setup_hours,
        row.repair_hours,
        row.saw_blade_exchange_hours,
        row.loss_stop_day_hours,
        row.loss_stop_night_hours,
        row.invisible_loss_stop_hours,
        row.planned_stop_hours,
        row.available_work_hours,
        row.work_hours,
        row.utilization_rate,
        row.work_rate,
        row.total_production_qty,
        row.efficiency_rate,
        row.equipment_record_time_1,
        row.equipment_record_time_2,
        row.metric_60,
        row.actual_record_cutting,
        row.cutting_record_variance,
        row.actual_record_chamfering,
        row.chamfering_record_variance,
        row.chamfering_setup_hours,
        row.forming_process_date,
        row.data_source,
        row.external_sync_key,
        row.remarks,
    )


def sync_parsed_rows_to_db(
    cursor,
    rows: list[ParsedCuttingIndicatorRow],
    *,
    source_file: str,
    dry_run: bool = False,
) -> ImportSyncResult:
    result = ImportSyncResult(parsed=len(rows))
    if dry_run:
        return result

    cursor.execute(
        "SELECT COUNT(1) FROM cutting_production_indicator WHERE source_file = %s",
        (source_file,),
    )
    result.deleted = int(cursor.fetchone()[0] or 0)
    if result.deleted:
        cursor.execute(
            "DELETE FROM cutting_production_indicator WHERE source_file = %s",
            (source_file,),
        )

    if not rows:
        return result

    placeholders = ", ".join(["%s"] * len(INSERT_COLUMNS))
    sql = f"INSERT INTO cutting_production_indicator ({', '.join(INSERT_COLUMNS)}) VALUES ({placeholders})"
    batch_params = [_row_to_insert_tuple(r) for r in rows]
    cursor.executemany(sql, batch_params)
    result.inserted = len(rows)
    return result


def sync_cutting_source_file(
    source_path: str,
    connection_factory,
    *,
    dry_run: bool = False,
    remarks_prefix: str = REMARKS_PREFIX_EXCEL,
) -> ImportSyncResult:
    """CSV/Excel 1 ファイルを cutting_production_indicator に全件置換同期"""
    source_label = os.path.basename(source_path)
    conn = connection_factory()
    conn.autocommit = False
    cursor = conn.cursor()
    try:
        rows, parse_errors, skipped = parse_rows_from_source(
            source_path,
            remarks_prefix=remarks_prefix,
        )
        result = sync_parsed_rows_to_db(cursor, rows, source_file=source_label, dry_run=dry_run)
        result.skipped_invalid = skipped
        result.errors.extend(parse_errors)
        if not dry_run:
            conn.commit()
        logger.info(
            "cutting_production_indicator 同期完了: parsed=%s inserted=%s deleted=%s path=%s",
            result.parsed,
            result.inserted,
            result.deleted,
            source_path,
        )
        return result
    except Exception:
        conn.rollback()
        raise
    finally:
        cursor.close()
        conn.close()
